"""Exportacao .xlsx real com openpyxl (secao 25): 4 abas formatadas."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import get_settings
from app.schemas.models import CriterionKind, RankedOffer, SearchResults

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MONEY = 'R$ #,##0.00'
_PCT = "0.0%"

_LABELS = {
    "melhor_custo_beneficio": "Melhor custo-benefício",
    "menor_preco": "Menor preço",
    "menor_preco_confiavel": "Menor preço confiável",
    "melhor_avaliado": "Melhor avaliado",
    "compra_mais_segura": "Compra mais segura",
    "melhor_importada": "Melhor opção importada",
    "opcao_economica": "Opção econômica",
    "opcao_intermediaria": "Opção intermediária",
    "opcao_premium": "Opção premium",
}


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 2)}"


def _autofit(ws: Worksheet, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(10, length + 2), max_width)


def _fmt_money(ws: Worksheet, row: int, cols: list[int]) -> None:
    for c in cols:
        ws.cell(row=row, column=c).number_format = _MONEY


def _reputation_text(r) -> str:
    return r.classification.replace("_", " ")


def build_workbook(results: SearchResults) -> bytes:
    wb = Workbook()

    # ---------------------------------------------------------- Aba 1
    ws = wb.active
    ws.title = "Comparativo"
    headers = [
        "Classificação", "Produto", "Marca", "Modelo", "Categoria", "Especificações",
        "Obrigatórios atendidos", "Obrigatórios não atendidos", "Desejáveis atendidos",
        "Loja", "Marketplace", "Vendedor", "Entrega por", "Origem", "Condição",
        "Preço normal", "Preço no Pix", "Parcelas s/ juros", "Valor da parcela",
        "Frete", "Impostos", "Taxas", "Cupom", "Regras do cupom", "Cashback",
        "Regras do cashback", "Preço total imediato", "Prazo (dias)", "Entrega confirmada",
        "Avaliação", "Qtd. avaliações", "Reputação loja", "Reputação vendedor", "Garantia",
        "Nota custo-benefício", "Vantagens", "Desvantagens", "Alertas", "Link", "Fonte",
        "Consultado em",
    ]
    ws.append(headers)

    all_rows: list[tuple[RankedOffer, bool]] = [(r, True) for r in results.offers] + [
        (r, False) for r in results.unvalidated_offers
    ]
    for idx, (r, validated) in enumerate(all_rows, start=2):
        o, b = r.offer, r.price_breakdown
        labels = ", ".join(_LABELS.get(lb, lb) for lb in r.labels) or (
            "" if validated else "Oferta não validada para o CEP informado"
        )
        specs = "; ".join(f"{k}: {v}" for k, v in o.specs.items())
        warranty = f"{o.warranty.kind.replace('_', ' ')} ({o.warranty.months} meses)"
        ws.append([
            labels, o.product_name, o.brand, o.model, o.category, specs,
            "; ".join(r.mandatory_met), "; ".join(r.mandatory_unmet), "; ".join(r.desirable_met),
            o.store, o.marketplace, o.seller_name, o.fulfilled_by, o.origin, o.condition,
            b.price, b.price_pix, o.installments_count if o.installments_interest_free else 0,
            o.installment_value, b.shipping, b.taxes, b.fees,
            b.coupon_discount, (o.coupon.rules if o.coupon else ""),
            b.cashback_later, (o.cashback.rules if o.cashback else ""),
            b.total_delivered, o.shipping_days, "Sim" if validated else "Não confirmada",
            o.reviews.average, o.reviews.count,
            _reputation_text(o.store_reputation), _reputation_text(o.seller_reputation),
            warranty, r.score, "; ".join(r.advantages), "; ".join(r.disadvantages),
            "; ".join(r.alerts), o.url, o.source,
            o.collected_at.strftime("%d/%m/%Y %H:%M"),
        ])
        _fmt_money(ws, idx, [16, 17, 19, 20, 21, 22, 23, 25, 27])
        link_cell = ws.cell(row=idx, column=39)
        if o.url:
            link_cell.hyperlink = o.url
            link_cell.font = Font(color="1D4ED8", underline="single")
        for col in (6, 36, 37, 38):
            ws.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    _style_header(ws, len(headers))
    _autofit(ws)

    # ---------------------------------------------------------- Aba 2
    ws2 = wb.create_sheet("Histórico de preços")
    h2 = [
        "Produto", "Marca", "Modelo", "Loja", "Vendedor", "Data", "Preço normal",
        "Preço no Pix", "Frete", "Impostos", "Taxas", "Preço total", "Cupom",
        "Cashback", "Disponibilidade", "CEP consultado",
    ]
    ws2.append(h2)
    row = 2
    cep_masked = results.cep.cep[:2] + "***-***"
    for r in results.offers:
        o = r.offer
        for point in r.history.series:
            ws2.append([
                o.product_name, o.brand, o.model, o.store, o.seller_name,
                point["date"], point["total_price"], point["total_price"],
                r.price_breakdown.shipping, r.price_breakdown.taxes, r.price_breakdown.fees,
                point["total_price"], r.price_breakdown.coupon_discount,
                r.price_breakdown.cashback_later, "Sim", cep_masked,
            ])
            _fmt_money(ws2, row, [7, 8, 9, 10, 11, 12, 13, 14])
            row += 1
    _style_header(ws2, len(h2))
    _autofit(ws2)

    # ---------------------------------------------------------- Aba 3
    ws3 = wb.create_sheet("Critérios da pesquisa")
    q = results.query
    weights = get_settings().ranking_weights
    kind_map = {
        CriterionKind.OBRIGATORIO: "Obrigatório",
        CriterionKind.DESEJAVEL: "Desejável",
        CriterionKind.INDIFERENTE: "Indiferente",
    }
    meta_rows = [
        ("Descrição original", q.original_text),
        ("CEP", results.cep.cep),
        ("Cidade", results.cep.city),
        ("Estado", results.cep.state),
        ("Preço máximo", f"R$ {q.max_price:,.2f}" if q.max_price else "não informado"),
        ("Importados permitidos", "Sim" if q.allow_imported else "Não"),
        ("Tolerância aplicada", f"{q.tolerance * 100:.0f}%"),
        ("Pesos do ranking", "; ".join(f"{k}: {v * 100:.0f}%" for k, v in weights.items())),
        ("Data e hora", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Critério", "Tipo"),
    ]
    for a, bvalue in meta_rows:
        ws3.append([a, bvalue])
    for crit in q.criteria:
        ws3.append([crit.label, kind_map[crit.kind]])
    for cell in ws3["A"]:
        cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 70

    # ---------------------------------------------------------- Aba 4
    ws4 = wb.create_sheet("Fontes e alertas")
    h4 = ["Fonte", "Tipo", "Simulada", "Status", "Ofertas", "Descartadas", "Observações", "Acesso em"]
    ws4.append(h4)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for s in results.sources:
        ws4.append([
            s.name, s.kind, "Sim" if s.simulated else "Não", s.status,
            s.offers_found, s.offers_discarded, s.message, now,
        ])
    ws4.append([])
    ws4.append(["Alertas por oferta"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True)
    for r in results.offers + results.unvalidated_offers:
        for alert in r.alerts:
            ws4.append([r.offer.product_name, r.offer.store, "", "", "", "", alert])
    _style_header(ws4, len(h4))
    _autofit(ws4, max_width=60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
