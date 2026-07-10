"""Fluxo completo de pesquisa no modo demonstracao, via API."""

import time

_LABEL_KEYS = {"melhor_custo_beneficio", "menor_preco", "compra_mais_segura"}


def _run_search(client, text: str, cep: str = "74000-000", **kw):
    interp = client.post("/api/interpret", json={"text": text, "cep": cep, **kw})
    assert interp.status_code == 200, interp.text
    body = interp.json()
    assert body["cep"]["city"] == "Goiânia"

    start = client.post("/api/search", json={"query": body["query"], "cep": cep})
    assert start.status_code == 200
    search_id = start.json()["search_id"]

    for _ in range(100):
        result = client.get(f"/api/search/{search_id}").json()
        if result["status"] != "executando":
            return search_id, result
        time.sleep(0.1)
    raise AssertionError("pesquisa não concluiu a tempo")


def test_fluxo_geladeira(client):
    search_id, result = _run_search(
        client, "Geladeira frost free, aproximadamente 450 litros, 220 V, nova, até R$ 5.000"
    )
    assert result["status"] == "concluida"
    assert result["amazon_consulted"] is True

    offers = result["offers"]
    assert len(offers) >= 3
    # a TF39 (390L, 110V) nao pode passar; a Panasonic sem CEP vai p/ nao validadas
    ids = {o["offer"]["offer_id"] for o in offers}
    assert "mgl-geladeira-electrolux-tf39-110v" not in ids
    unvalidated_ids = {o["offer"]["offer_id"] for o in result["unvalidated_offers"]}
    assert "amz-geladeira-panasonic-bb53-semcep" in unvalidated_ids

    assert _LABEL_KEYS.issubset(result["highlights"].keys())
    # menor preco e o Samsung do vendedor ruim; menor preco confiavel deve ser outro
    assert result["highlights"]["menor_preco"] == "mgl-geladeira-samsung-rt46-barata"
    assert result["highlights"]["menor_preco_confiavel"] != "mgl-geladeira-samsung-rt46-barata"

    # historico demo disponivel
    assert any(o["history"]["available"] for o in offers)
    # ofertas ordenadas por nota
    scores = [o["score"] for o in offers]
    assert scores == sorted(scores, reverse=True)


def test_fluxo_importado_bloqueado(client):
    _, result = _run_search(
        client, "Notebook com 16 GB de RAM, SSD de 512 GB", allow_imported=False
    )
    assert all(o["offer"]["origin"] == "nacional" for o in result["offers"])


def test_fluxo_importado_permitido_com_alerta(client):
    _, result = _run_search(client, "Notebook com 16 GB de RAM, SSD de 512 GB")
    imported = [o for o in result["offers"] if o["offer"]["origin"] == "importado"]
    assert imported, "esperava oferta importada no ranking"
    assert any("importação" in a or "importado" in a for a in imported[0]["alerts"])
    assert "melhor_importada" in result["highlights"]


def test_export_excel(client):
    search_id, result = _run_search(client, "Ar-condicionado inverter 12.000 BTUs 220 V")
    resp = client.get(f"/api/search/{search_id}/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    assert resp.content[:2] == b"PK"  # xlsx real (zip), nao CSV renomeado

    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == [
        "Comparativo", "Histórico de preços", "Critérios da pesquisa", "Fontes e alertas",
    ]
    ws = wb["Comparativo"]
    assert ws.max_row >= 3  # cabecalho + 2 ofertas de ar-condicionado


def test_cep_invalido_bloqueia(client):
    resp = client.post("/api/interpret", json={"text": "geladeira", "cep": "00000-000"})
    assert resp.status_code == 422


def test_health_e_fontes(client):
    health = client.get("/health").json()
    assert health["status"] == "ok"
    assert health["mode"] == "demo"
    sources = client.get("/api/sources").json()
    names = {s["name"] for s in sources}
    assert "Amazon Brasil" in names
    assert len(names) >= 3
